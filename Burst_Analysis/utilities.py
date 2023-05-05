import numpy as np
import pandas as pd
import warnings
import datetime

warnings.filterwarnings('ignore')
from statistics import multimode

from statistics import mode
import itertools
from openpyxl import load_workbook

""" 
    This file contains all the helper functions needed to calculate the burst variables 
    and inter-reward interval variables.
"""


def cleanup(filepath):
    """ This function is used to cleanup the excel output file for further analysis

    Args:
        filepath (_type_): str

    Returns:
        _type_: pd.DataFrame
    """
    # print(filepath)
    # import data and transpose
    df_raw = pd.read_excel(filepath, header=None).T

    # modify the header
    new_header = df_raw.iloc[0]  # grab the first row for the header
    df = df_raw[1:]  # take the data except the header row
    df.columns = new_header
    df.reset_index(drop=True, inplace=True)

    # add new column for visualization 
    df['Start Datetime'] = pd.to_datetime(df['Start Date'] + ' ' + df['Start Time'])
    df['End Datetime'] = pd.to_datetime(df['End Date'] + ' ' + df['End Time'])
    df['Timeout'] = (df['Active Lever Presses'] - df['Reward']).astype('int32')

    # drop unnecessary columns
    df.drop(['Start Date', 'Start Time', 'End Date', 'End Time'], axis=1, inplace=True)

    # change data types
    cols = df.columns.tolist()
    for col in cols[:-1]:
        name = col.lower()
        if ('active' in name) or ('reward' in name) or ('timeout' in name):
            df[col] = df[col].astype('int32')
            df[col] = df[col].replace(0, np.nan)
        else:
            pass

    # drop columns which all values are Nan
    df.dropna(how='all', axis=1, inplace=True)

    # reorder the columns
    new_columns = df.columns.tolist()
    new_columns.insert(0, new_columns.pop(new_columns.index('Subject')))
    new_columns.insert(1, new_columns.pop(new_columns.index('Start Datetime')))
    new_columns.insert(2, new_columns.pop(new_columns.index('End Datetime')))
    idx = new_columns.index('Reward') + 1
    new_columns.insert(idx, new_columns.pop(new_columns.index('Timeout')))

    # fill the nan with 0
    df = df[new_columns]
    df.fillna(0, inplace=True)
    # print('CLEANING COMPLETED')
    # output the result dataframe
    return df


def get_mode(lst):
    """ This function returns the mode of the input list

    Args:
        lst (_type_): list of int

    Returns:
        _type_: int
    """
    if len(lst) == 0:
        return np.NaN
    return mode(lst)


def get_bursts(lst, interval=90):
    """ This function retrieves the "bursts" (cluster of rewards happened within 2 mins) from rewards

    Args:
        lst (_type_): list of int

    Returns:
        _type_: 2d list
    """
    allBursts = []
    i = 0
    while i < len(lst):
        oneBurst = []
        limit = lst[i] + interval
        j = i
        while j < len(lst) and lst[j] <= limit:
            oneBurst.append(lst[j])
            limit = lst[j] + interval
            j += 1
        allBursts.append(oneBurst)
        i = j
    return allBursts


def get_mean_num_rewards(lst):
    """ This function returns the mean number of rewards out of all the bursts

    Args:
        lst (_type_): list of int

    Returns:
        _type_: int
    """
    bursts = [i for i in lst if len(i) > 1]
    total_rewards = len(list(itertools.chain.from_iterable(bursts)))
    total_bursts = len(bursts)

    if total_bursts == 0:
        return 0

    return round(total_rewards / total_bursts, 2)


def get_burst_rewards_pct(lst):
    """ This function returns the percentage of rewards that fall in burst out of all the rewards

    Args:
        lst (_type_): 2d list

    Returns:
        _type_: int
    """
    numRewards = len(list(itertools.chain.from_iterable(lst)))
    if numRewards == 0:
        return np.NaN
    bursts = [i for i in lst if len(i) > 1]
    numBurstRewards = len(list(itertools.chain.from_iterable(bursts)))
    return round(numBurstRewards / numRewards * 100, 2)


def get_max_burst(lst):
    """ This function returns the max number of rewards amoung all the bursts

    Args:
        lst (_type_): 2d list 

    Returns:
        _type_: int
    """
    bursts = [i for i in lst if len(i) > 1]
    if len(bursts) == 0:
        return 0
    return max([len(i) for i in bursts])


def calculations_single(df):
    """ This function returns the dataframe of all the calculated variables (original version)

    Args:
        df (_type_): pd.DataFrame

    Returns:
        _type_: pd.DataFrame
    """
    # combine all reward timestamp into a column of lists
    filtered_cols = ['Subject'] + [col for col in df.columns if 'Reward ' in col]
    df_reward = df[filtered_cols].sort_values('Subject').reset_index().drop('index', axis=1)
    df_reward['allRewards'] = df_reward.iloc[:, 1:].values.tolist()

    # get the filtered df
    dff = df_reward[['Subject', 'allRewards']]

    # retrieve the inter-reward intervals, filtering out negatives
    dff['Intervals'] = dff['allRewards'].apply(lambda lst: [j - i for i, j in zip(lst[:-1], lst[1:])])
    dff['cleanedIntervals'] = dff['Intervals'].apply(lambda lst: [val for val in lst if val > 0])

    # calculate needed traits related to the inter-reward intervals
    dff['meanInterval'] = dff['cleanedIntervals'].apply(lambda x: round(np.mean(x), 2))
    dff['stdInterval'] = dff['cleanedIntervals'].apply(lambda x: round(np.std(x), 2))
    dff['modeInterval'] = dff['cleanedIntervals'].apply(get_mode)

    # filtering the rewards (zero values)
    dff['cleanedRewards'] = dff['allRewards'].apply(lambda lst: [val for val in lst if val > 0])
    dff['totalRewards'] = dff['cleanedRewards'].apply(lambda x: len(x))

    # retrieve the "bursts" (cluster of rewards happened within 2 mins) from rewards 
    dff['rawBurst'] = dff['cleanedRewards'].apply(get_bursts)
    dff['numBurst'] = dff['rawBurst'].apply(lambda x: len([i for i in x if len(i) > 1]))

    # get the mean number of rewards across all the bursts
    dff['meanNumRewards'] = dff['rawBurst'].apply(get_mean_num_rewards)

    # get the percentage of rewards that fall in burst out of all the rewards
    dff['pctRewards'] = dff['rawBurst'].apply(get_burst_rewards_pct)

    # get the maximum number of rewards contain in a single burst in one session
    dff['maxBurst'] = dff['rawBurst'].apply(get_max_burst)

    # select needed columns for output df
    output_cols = ['Subject', 'meanInterval', 'stdInterval', 'modeInterval', 'meanNumRewards', 'numBurst', 'maxBurst',
                   'pctRewards']
    dff_out = dff[output_cols]

    return dff_out


def get_sheetnames_xlsx(file_name):
    """This function returns the worksheet names from an given Excel workbook

    Args:
        file_name (_type_): str

    Returns:
        _type_: list of strings
    """
    wb = load_workbook(file_name, read_only=True, keep_links=False)
    return wb.sheetnames


def filtered_reward(df):
    """This function returns a df that contains reward latency, all rewards, intervals, and cleaned intervals

    Args:
        df (_type_): pd.DataFrame

    Returns:
        _type_: pd.DataFrame
    """
    filtered_cols = ['Subject'] + [col for col in df.columns if 'Reward ' in col]
    df_reward = df[filtered_cols].sort_values('Subject').reset_index().drop('index', axis=1)
    df_reward['allRewards'] = df_reward.iloc[:, 1:].values.tolist()
    df_reward['Latency'] = df_reward['Reward 1']
    df_filtered = df_reward[['Subject', 'Latency', 'allRewards']]
    df_filtered['Intervals'] = df_filtered['allRewards'].apply(lambda lst: [j - i for i, j in zip(lst[:-1], lst[1:])])
    df_filtered['cleanedIntervals'] = df_filtered['Intervals'].apply(lambda lst: [val for val in lst if val > 0][:-1])

    return df_filtered


def unify_date(timestamp):
    """unifies the date of the timestamp for 'Start' and 'End' column in the timeline_data

    Args:
        timestamp (_type_): _description_

    Returns:
        _type_: Timestamp
    """
    date = pd.Timestamp("2000-01-01")
    time = timestamp.time()
    return pd.to_datetime(datetime.datetime.combine(date, time))


def unstack_data(df):
    rows = []
    rfids = [i for i in df.rfid.unique()]

    for rfid in rfids:
        temp = df[df.rfid==rfid]
        temp = temp.reset_index(drop=True)
        temp.sort_values(['trial_id'],inplace=True)
        temp.set_index('trial_id',inplace=True)
        temp_row = temp.iloc[:,1:].unstack().to_frame().sort_index(level=1).T
        temp_row.insert(loc=0, column='rfid', value=temp.rfid.unique())
        temp_row = temp_row.reset_index(drop=True)
        rows.append(temp_row)
        
    return rows


def deserialize_data(ts):
    if pd.isnull(ts):
        return []
    return [float(i) for i in ts.split()]