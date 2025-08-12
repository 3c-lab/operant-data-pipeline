import openpyxl
import os
import re
from datetime import datetime

FIRST_SESSION_COL = 5

def sort_columns_by_date(wb, source_sheet_name="Behavior Data", sorted_sheet_name="Sorted Data"):
    """
    Sort session columns (D..end) by the date in row=1 (ascending).
    Remove the old (unsorted) sheet and rename the new sheet to 'Behavior Data'.
    """
    if source_sheet_name not in wb.sheetnames:
        print(f"Sheet '{source_sheet_name}' not found! Nothing to sort.")
        return


    source_ws = wb[source_sheet_name]


    if sorted_sheet_name in wb.sheetnames:
        del wb[sorted_sheet_name]
    sorted_ws = wb.create_sheet(title=sorted_sheet_name)


    max_row = source_ws.max_row
    max_col = source_ws.max_column


    # Because columns 1..3 => Rat, RFID, Drug Group, session columns start at 4
    columns_info = []
    for col_idx in range(FIRST_SESSION_COL, max_col + 1):
        date_val = source_ws.cell(row=1, column=col_idx).value
        if date_val:
            if isinstance(date_val, datetime):
                parsed_date = date_val
            else:
                try:
                    parsed_date = datetime.strptime(str(date_val), "%m/%d/%Y")
                except ValueError:
                    try:
                        parsed_date = datetime.strptime(str(date_val), "%Y-%m-%d")
                    except ValueError:
                        parsed_date = datetime.max
            columns_info.append((col_idx, parsed_date, date_val))
        else:
            columns_info.append((col_idx, datetime.max, None))


    # Sort ascending by parsed_date
    columns_info.sort(key=lambda x: x[1])


    # Copy columns A,B,C, D
    for r in range(1, max_row + 1):
        for c in range(1, FIRST_SESSION_COL):
            sorted_ws.cell(row=r, column=c, value=source_ws.cell(row=r, column=c).value)


    # Copy sorted columns to new sheet, starting at col=4
    for idx, (old_col_idx, _, _) in enumerate(columns_info):
        new_col_idx = FIRST_SESSION_COL + idx
        for r in range(1, max_row + 1):
            sorted_ws.cell(row=r, column=new_col_idx,
                           value=source_ws.cell(row=r, column=old_col_idx).value)


    wb.remove(source_ws)
    sorted_ws.title = source_sheet_name
    print(f"Session columns sorted by date. Old sheet removed. New sheet is '{source_sheet_name}'.")




def find_or_create_section(ws, heading_name):
    """
    Check if heading_name exists in col A. If not, append at bottom. Return row index.
    """
    max_row = ws.max_row
    for rr in range(1, max_row + 1):
        if ws.cell(rr, 1).value == heading_name:
            return rr
    new_row = max_row + 1
    ws.cell(row=new_row, column=1, value=heading_name)
    return new_row

def main():
    # -------------------------------------------------------------------------
    # 0) SET THE COHORT NUMBER
    # -------------------------------------------------------------------------
    cohort_number = 29  # <--- ADJUST HERE
    cohort_str = f"C{cohort_number:02d}"  # e.g., "C21", "C07"


    # Output file path
    behavior_sheet_path = (
        rf"C:\Users\georg\George Lab Dropbox\George_Lab\Team GWAS\1. Cocaine GWAS\Cocaine Behavior Sheet Auto"
        rf"\Cocaine_{cohort_str}_behavior_sheet.xlsx"
    )


    # -------------------------------------------------------------------------
    # 1) FOLDER LOCATIONS
    # -------------------------------------------------------------------------
    # "Cocaine Cohort Information" folder
    cohort_folder = r"C:\Users\georg\George Lab Dropbox\George_Lab\Team GWAS\1. Cocaine GWAS\Cocaine Cohort Information"
    cohort_file_name = f"Cocaine {cohort_str} Updated Cohort Information.xlsx"
    cohort_file_path = os.path.join(cohort_folder, cohort_file_name)


    # "Cocaine Daily Issues" folder
    daily_issues_folder = r"C:\Users\georg\George Lab Dropbox\George_Lab\Team GWAS\1. Cocaine GWAS\Cocaine Daily Issues"
    daily_issues_file_name = f"Coc_{cohort_str}_Issues.xlsx"
    daily_issues_file_path = os.path.join(daily_issues_folder, daily_issues_file_name)


    # "excel_output_files" => "COCAINE"
    base_excel_output_folder = r"C:\Users\georg\George Lab Dropbox\George_Lab\Experiments\DataStream\DataSource\excel_output_files"
    excel_output_files = os.path.join(base_excel_output_folder, "COCAINE")


    import openpyxl


    # -------------------------------------------------------------------------
    # 2) OPEN OR CREATE THE WORKBOOK & CLEAR SHEET
    # -------------------------------------------------------------------------
    if os.path.exists(behavior_sheet_path):
        behavior_wb = openpyxl.load_workbook(behavior_sheet_path)
        if "Behavior Data" in behavior_wb.sheetnames:
            del behavior_wb["Behavior Data"]
        behavior_ws = behavior_wb.create_sheet("Behavior Data")
    else:
        behavior_wb = openpyxl.Workbook()
        behavior_ws = behavior_wb.active
        behavior_ws.title = "Behavior Data"


    # -------------------------------------------------------------------------
    # 3) Minimal headers
    # -------------------------------------------------------------------------
    if not behavior_ws["A1"].value:
        behavior_ws["A1"] = "Date"
    if not behavior_ws["A2"].value:
        behavior_ws["A2"] = "Session Name"


    # -------------------------------------------------------------------------
    # 4) LOAD COHORT WORKBOOK
    # -------------------------------------------------------------------------
    timeline_ws = None
    exit_tab_ws = None
    if os.path.exists(cohort_file_path):
        cohort_wb = openpyxl.load_workbook(cohort_file_path)
        timeline_ws = cohort_wb["Timeline"]
        exit_tab_ws = cohort_wb["Exit Tab"]
    else:
        print(f"Warning: Cohort info file not found at: {cohort_file_path}")


    # -------------------------------------------------------------------------
    # 5) REWARDS => "A5"
    # -------------------------------------------------------------------------
    behavior_ws["A5"] = "Rewards"


    # RAT, RFID, Drug Group
    if timeline_ws:
        behavior_ws["A6"] = "Rat"
        behavior_ws["B6"] = "RFID"
        behavior_ws["C6"] = "Drug Group"
        behavior_ws["D6"] = "Experiment Group" #Experiment Group


        rfid_col_idx = None
        drug_group_col_idx = None
        exp_group_col_idx  = None #Experiment Group


        header_row = list(timeline_ws[1])
        for cidx, cell in enumerate(header_row, start=1):
            if cell.value == "RFID":
                rfid_col_idx = cidx
            if cell.value == "Drug Group":
                drug_group_col_idx = cidx
            if cell.value == "Experiment Group": #Experiment Group
                exp_group_col_idx = cidx


        row_offset = 1
        for row_data in timeline_ws.iter_rows(min_row=2):
            rat_val = row_data[0].value
            rfid_val = None
            if rfid_col_idx is not None and (rfid_col_idx - 1) < len(row_data):
                rfid_val = row_data[rfid_col_idx - 1].value
            drug_val = None
            if drug_group_col_idx is not None and (drug_group_col_idx - 1) < len(row_data):
                drug_val = row_data[drug_group_col_idx - 1].value
            exp_val = None                                                                            #Experiment Group
            if exp_group_col_idx is not None and (exp_group_col_idx - 1) < len(row_data):
                exp_val = row_data[exp_group_col_idx - 1].value

            new_row = 6 + row_offset
            behavior_ws.cell(row=new_row, column=1, value=rat_val)
            behavior_ws.cell(row=new_row, column=2, value=rfid_val)
            behavior_ws.cell(row=new_row, column=3, value=drug_val)
            behavior_ws.cell(row=new_row, column=4, value=exp_val)       #Experiment Group
            row_offset += 1
    else:
        timeline_ws = None


    # Gather RAT/RFID/Drug
    original_rat_rfid_dg = []
    row_idx = 7
    while True:
        rat_val = behavior_ws.cell(row=row_idx, column=1).value
        if rat_val is None:
            break
        if rat_val not in ["Rat", "Active Lever Presses", "Inactive Lever Presses", 
                           "breakpoint", "Total Shocks", "Reward # Got Shock 1"]:
            rfid_val = behavior_ws.cell(row=row_idx, column=2).value
            drug_val = behavior_ws.cell(row=row_idx, column=3).value
            exp_val  = behavior_ws.cell(row=row_idx, column=4).value
            original_rat_rfid_dg.append((rat_val, rfid_val, drug_val, exp_val))
        row_idx += 1


    # Insert Exit Day/Code/Notes
    last_col_for_exit = 4
    behavior_ws.cell(row=6, column=last_col_for_exit + 1, value="Exit Day")
    behavior_ws.cell(row=6, column=last_col_for_exit + 2, value="Exit Code")
    behavior_ws.cell(row=6, column=last_col_for_exit + 3, value="Exit Notes")
    behavior_ws.cell(row=6, column=last_col_for_exit + 4, value="Last Good Session") #Updated


    exit_day_col = last_col_for_exit + 1
    exit_code_col = last_col_for_exit + 2
    exit_notes_col = last_col_for_exit + 3 
    last_good_col = last_col_for_exit + 4  #Updated

    # EXIT TAB
    exit_data_map = {}
    if exit_tab_ws:
        exit_tab_headers = {}
        for cidx, cell in enumerate(exit_tab_ws[1], start=1):
            val = cell.value
            header_str = str(val).strip() if val else ""
            exit_tab_headers[header_str] = cidx


        rat_col_idx_exit = exit_tab_headers.get("RAT")
        exit_day_idx = exit_tab_headers.get("Exit Day")
        exit_code_idx = exit_tab_headers.get("Exit Code")
        exit_notes_idx = exit_tab_headers.get("Exit Notes")
        last_good_idx = exit_tab_headers.get("Last Good Session") #Updated


        if rat_col_idx_exit and exit_day_idx and exit_code_idx and exit_notes_idx:
            for r_i in range(2, exit_tab_ws.max_row + 1):
                rat_val_exit = exit_tab_ws.cell(row=r_i, column=rat_col_idx_exit).value
                if rat_val_exit:
                    day_v = exit_tab_ws.cell(row=r_i, column=exit_day_idx).value
                    code_v = exit_tab_ws.cell(row=r_i, column=exit_code_idx).value
                    notes_v = exit_tab_ws.cell(row=r_i, column=exit_notes_idx).value
                    last_good_v = exit_tab_ws.cell(row=r_i, column=last_good_idx ).value
                    exit_data_map[str(rat_val_exit).strip()] = (day_v, code_v, notes_v, last_good_v) #Updated
        else:
            print("Warning: Some columns (RAT/Exit Day/Exit Code/Exit Notes) missing in the Exit Tab.")
    else:
        print("Warning: Exit Tab not loaded properly.")


    # Subfolders
    cohort_pattern = re.compile(rf"{cohort_str}")
    session_pattern = re.compile(r"HSCOC(\w+)_output\.xlsx$")
    subfolders = ["LGA", "SHA", "PR", "SHOCK"]
    session_names = set()


    for folder in subfolders:
        sf_path = os.path.join(excel_output_files, folder)
        if not os.path.exists(sf_path):
            continue
        for f_name in os.listdir(sf_path):
            if (f_name.startswith("MTF134") or f_name.startswith("BSB273")) and cohort_pattern.search(f_name):
                match = session_pattern.search(f_name)
                if match:
                    session_names.add(match.group(1))


    # Insert session columns
    existing_sessions = set()
    max_col_now = behavior_ws.max_column
    for col_i in range(FIRST_SESSION_COL, max_col_now + 1):    
        val = behavior_ws.cell(row=2, column=col_i).value
        if val:
            existing_sessions.add(val)


    start_col = max_col_now + 1 if max_col_now >= FIRST_SESSION_COL else FIRST_SESSION_COL
    for s_name in sorted(session_names):
        if s_name not in existing_sessions:
            behavior_ws.cell(row=2, column=start_col, value=s_name)
            start_col += 1


    max_col_now = behavior_ws.max_column
    all_sessions_in_sheet = [behavior_ws.cell(row=2, column=c).value for c in range(FIRST_SESSION_COL, max_col_now + 1)]


    # Insert 'Start Date'
    for s_name in all_sessions_in_sheet:
        if not s_name:
            continue
        for folder in subfolders:
            sf_path = os.path.join(excel_output_files, folder)
            if not os.path.exists(sf_path):
                continue


            for f_name in os.listdir(sf_path):
                if (f_name.startswith("MTF134") or f_name.startswith("BSB273")) \
                   and s_name in f_name and f_name.endswith("_output.xlsx") and cohort_pattern.search(f_name):


                    session_file_path = os.path.join(sf_path, f_name)
                    session_wb = openpyxl.load_workbook(session_file_path)
                    session_ws = session_wb.active


                    for row_cells in session_ws.iter_rows(min_col=1, max_col=1):
                        if row_cells[0].value == "Start Date":
                            date_cell = session_ws.cell(row=row_cells[0].row, column=2)
                            if date_cell.value:
                                for cc in range(FIRST_SESSION_COL, behavior_ws.max_column + 1):
                                    if behavior_ws.cell(row=2, column=cc).value == s_name:
                                        behavior_ws.cell(row=1, column=cc, value=date_cell.value)
                                        break


    # -------------------------------------------------------------------------
    # REWARDS
    #   shock => "Total Reward"
    # -------------------------------------------------------------------------
    for s_name in all_sessions_in_sheet:
        if not s_name:
            continue
        for folder in subfolders:
            sf_path = os.path.join(excel_output_files, folder)
            if not os.path.exists(sf_path):
                continue


            reward_label = "Total Reward" if folder == "SHOCK" else "Reward"


            for f_name in os.listdir(sf_path):
                if (f_name.startswith("MTF134") or f_name.startswith("BSB273")) \
                   and s_name in f_name and f_name.endswith("_output.xlsx") \
                   and cohort_pattern.search(f_name):


                    session_file_path = os.path.join(sf_path, f_name)
                    session_wb = openpyxl.load_workbook(session_file_path)
                    session_ws = session_wb.active


                    reward_row = None
                    subject_row = None
                    for row_cells in session_ws.iter_rows(min_col=1, max_col=1):
                        if row_cells[0].value == reward_label:
                            reward_row = row_cells[0].row
                        if row_cells[0].value == "Subject":
                            subject_row = row_cells[0].row


                    if reward_row and subject_row:
                        reward_data = [cell.value for cell in session_ws[reward_row]]
                        subject_data = [cell.value for cell in session_ws[subject_row]]


                        for rat_idx, rat_id in enumerate(subject_data[1:], start=1):
                            if rat_id in [r[0] for r in original_rat_rfid_dg]:
                                for row_b in range(7, behavior_ws.max_row + 1):
                                    if behavior_ws.cell(row=row_b, column=1).value == rat_id:
                                        for c_b in range(FIRST_SESSION_COL, behavior_ws.max_column + 1):
                                            if behavior_ws.cell(row=2, column=c_b).value == s_name:
                                                behavior_ws.cell(row=row_b, column=c_b, value=reward_data[rat_idx])
                                                break


    # -------------------------------------------------------------------------
    # ACTIVE LEVER PRESSES
    #   shock => "Total Active Lever Presses"
    # -------------------------------------------------------------------------
    active_row = find_or_create_section(behavior_ws, "Active Lever Presses")
    active_header_row = active_row + 1
    if behavior_ws.cell(active_header_row, 1).value != "Rat":
        behavior_ws.cell(active_header_row, 1, "Rat")
        behavior_ws.cell(active_header_row, 2, "RFID")
        behavior_ws.cell(active_header_row, 3, "Drug Group")
        behavior_ws.cell(active_header_row, 4, "Experiment Group")          #Experiment Group
        roff = 1
        for (rat_val, rfid_val, drug_val, exp_val) in original_rat_rfid_dg:
            behavior_ws.cell(active_header_row + roff, 1, rat_val)
            behavior_ws.cell(active_header_row + roff, 2, rfid_val)
            behavior_ws.cell(active_header_row + roff, 3, drug_val)
            behavior_ws.cell(active_header_row + roff, 4, exp_val)        #Experiment Group
            roff += 1


    active_rat_start = active_header_row + 1
    active_rat_ids = []
    for rr in range(active_rat_start, behavior_ws.max_row + 1):
        val = behavior_ws.cell(rr, 1).value
        if val and val not in ["Rat", "Active Lever Presses", "Inactive Lever Presses", 
                               "breakpoint", "Total Shocks", "Reward # Got Shock 1"]:
            active_rat_ids.append(val)


    for s_name in all_sessions_in_sheet:
        if not s_name:
            continue
        for folder in subfolders:
            sf_path = os.path.join(excel_output_files, folder)
            if not os.path.exists(sf_path):
                continue


            active_label = "Total Active Lever Presses" if folder == "SHOCK" else "Active Lever Presses"


            for f_name in os.listdir(sf_path):
                if (f_name.startswith("MTF134") or f_name.startswith("BSB273")) \
                   and s_name in f_name and f_name.endswith("_output.xlsx") \
                   and cohort_pattern.search(f_name):


                    session_file_path = os.path.join(sf_path, f_name)
                    session_wb = openpyxl.load_workbook(session_file_path)
                    session_ws = session_wb.active


                    active_press_row = None
                    subject_row_in_file = None
                    for row_cells in session_ws.iter_rows(min_col=1, max_col=1):
                        if row_cells[0].value == active_label:
                            active_press_row = row_cells[0].row
                        if row_cells[0].value == "Subject":
                            subject_row_in_file = row_cells[0].row


                    if active_press_row and subject_row_in_file:
                        active_data = [cell.value for cell in session_ws[active_press_row]]
                        subject_data = [cell.value for cell in session_ws[subject_row_in_file]]
                        for rat_idx, rat_id in enumerate(subject_data[1:], start=1):
                            if rat_id in active_rat_ids:
                                for row_b in range(active_rat_start, behavior_ws.max_row + 1):
                                    if behavior_ws.cell(row_b, 1).value == rat_id:
                                        for c_b in range(FIRST_SESSION_COL, behavior_ws.max_column + 1):
                                            if behavior_ws.cell(row=2, column=c_b).value == s_name:
                                                behavior_ws.cell(row=row_b, column=c_b, value=active_data[rat_idx])
                                                break


    # -------------------------------------------------------------------------
    # INACTIVE LEVER PRESSES
    #   shock => "Total Inactive Lever Presses"
    # -------------------------------------------------------------------------
    inactive_row = find_or_create_section(behavior_ws, "Inactive Lever Presses")
    inactive_header_row = inactive_row + 1
    if behavior_ws.cell(inactive_header_row, 1).value != "Rat":
        behavior_ws.cell(inactive_header_row, 1, "Rat")
        behavior_ws.cell(inactive_header_row, 2, "RFID")
        behavior_ws.cell(inactive_header_row, 3, "Drug Group")
        behavior_ws.cell(inactive_header_row, 4, "Experiment Group") #Experiment Group
        roff = 1
        for (rat_val, rfid_val, drug_val, exp_val) in original_rat_rfid_dg:
            behavior_ws.cell(inactive_header_row + roff, 1, rat_val)
            behavior_ws.cell(inactive_header_row + roff, 2, rfid_val)
            behavior_ws.cell(inactive_header_row + roff, 3, drug_val)
            behavior_ws.cell(inactive_header_row + roff, 4, exp_val) #Experiment Group
            roff += 1


    inactive_rat_start = inactive_header_row + 1
    inactive_rat_ids = []
    for rr in range(inactive_rat_start, behavior_ws.max_row + 1):
        val = behavior_ws.cell(rr, 1).value
        if val and val not in ["Rat", "Active Lever Presses", "Inactive Lever Presses", 
                               "breakpoint", "Total Shocks", "Reward # Got Shock 1"]:
            inactive_rat_ids.append(val)


    for s_name in all_sessions_in_sheet:
        if not s_name:
            continue
        for folder in subfolders:
            sf_path = os.path.join(excel_output_files, folder)
            if not os.path.exists(sf_path):
                continue


            inact_label = "Total Inactive Lever Presses" if folder == "SHOCK" else "Inactive Lever Presses"


            for f_name in os.listdir(sf_path):
                if (f_name.startswith("MTF134") or f_name.startswith("BSB273")) \
                   and s_name in f_name and f_name.endswith("_output.xlsx") \
                   and cohort_pattern.search(f_name):


                    session_file_path = os.path.join(sf_path, f_name)
                    session_wb = openpyxl.load_workbook(session_file_path)
                    session_ws = session_wb.active


                    inact_press_row = None
                    subject_row_in_file = None
                    for row_cells in session_ws.iter_rows(min_col=1, max_col=1):
                        if row_cells[0].value == inact_label:
                            inact_press_row = row_cells[0].row
                        if row_cells[0].value == "Subject":
                            subject_row_in_file = row_cells[0].row


                    if inact_press_row and subject_row_in_file:
                        inact_data = [cell.value for cell in session_ws[inact_press_row]]
                        subject_data = [cell.value for cell in session_ws[subject_row_in_file]]
                        for rat_idx, rat_id in enumerate(subject_data[1:], start=1):
                            if rat_id in inactive_rat_ids:
                                for row_b in range(inactive_rat_start, behavior_ws.max_row + 1):
                                    if behavior_ws.cell(row_b, 1).value == rat_id:
                                        for c_b in range(FIRST_SESSION_COL, behavior_ws.max_column + 1):
                                            if behavior_ws.cell(row=2, column=c_b).value == s_name:
                                                behavior_ws.cell(row=row_b, column=c_b, value=inact_data[rat_idx])
                                                break


    # -------------------------------------------------------------------------
    # BREAKPOINT
    # -------------------------------------------------------------------------
    breakpoint_row = find_or_create_section(behavior_ws, "breakpoint")
    breakpoint_header_row = breakpoint_row + 1
    if behavior_ws.cell(breakpoint_header_row, 1).value != "Rat":
        behavior_ws.cell(breakpoint_header_row, 1, "Rat")
        behavior_ws.cell(breakpoint_header_row, 2, "RFID")
        behavior_ws.cell(breakpoint_header_row, 3, "Drug Group")
        behavior_ws.cell(breakpoint_header_row, 4, "Experiment Group")
        roff = 1
        for (rat_val, rfid_val, drug_val, exp_val) in original_rat_rfid_dg:
            behavior_ws.cell(breakpoint_header_row + roff, 1, rat_val)
            behavior_ws.cell(breakpoint_header_row + roff, 2, rfid_val)
            behavior_ws.cell(breakpoint_header_row + roff, 3, drug_val)
            behavior_ws.cell(breakpoint_header_row + roff, 4, exp_val)
            roff += 1


    bp_rat_start = breakpoint_header_row + 1
    bp_rat_ids = []
    for rr in range(bp_rat_start, behavior_ws.max_row + 1):
        val = behavior_ws.cell(rr, 1).value
        if val and val not in ["Rat", "Active Lever Presses", "Inactive Lever Presses", 
                               "breakpoint", "Total Shocks", "Reward # Got Shock 1"]:
            bp_rat_ids.append(val)


    pr_folder_path = os.path.join(excel_output_files, "PR")
    if os.path.exists(pr_folder_path):
        for s_name in all_sessions_in_sheet:
            if not s_name:
                continue
            for f_name in os.listdir(pr_folder_path):
                if (f_name.startswith("MTF134") or f_name.startswith("BSB273")) \
                   and s_name in f_name and f_name.endswith("_output.xlsx") \
                   and cohort_pattern.search(f_name):


                    session_file_path = os.path.join(pr_folder_path, f_name)
                    session_wb = openpyxl.load_workbook(session_file_path)
                    session_ws = session_wb.active


                    last_ratio_row = None
                    subject_row_in_file = None
                    for row_cells in session_ws.iter_rows(min_col=1, max_col=1):
                        if row_cells[0].value == "Last Ratio":
                            last_ratio_row = row_cells[0].row
                        if row_cells[0].value == "Subject":
                            subject_row_in_file = row_cells[0].row


                    if last_ratio_row and subject_row_in_file:
                        last_ratio_data = [cell.value for cell in session_ws[last_ratio_row]]
                        subject_data = [cell.value for cell in session_ws[subject_row_in_file]]
                        for rat_idx, rat_id in enumerate(subject_data[1:], start=1):
                            if rat_id in bp_rat_ids:
                                for row_b in range(bp_rat_start, behavior_ws.max_row + 1):
                                    if behavior_ws.cell(row_b, 1).value == rat_id:
                                        for c_b in range(FIRST_SESSION_COL, behavior_ws.max_column + 1):
                                            if behavior_ws.cell(row=2, column=c_b).value == s_name:
                                                behavior_ws.cell(row=row_b, column=c_b, value=last_ratio_data[rat_idx])
                                                break


    # -------------------------------------------------------------------------
    # TOTAL SHOCKS
    # -------------------------------------------------------------------------
    total_shock_row = find_or_create_section(behavior_ws, "Total Shocks")
    total_shock_header_row = total_shock_row + 1
    if behavior_ws.cell(total_shock_header_row, 1).value != "Rat":
        behavior_ws.cell(total_shock_header_row, 1, "Rat")
        behavior_ws.cell(total_shock_header_row, 2, "RFID")
        behavior_ws.cell(total_shock_header_row, 3, "Drug Group")
        behavior_ws.cell(total_shock_header_row, 4, "Experiment Group") #Experiment Group
        roff = 1
        for (rat_val, rfid_val, drug_val, exp_val) in original_rat_rfid_dg:
            behavior_ws.cell(total_shock_header_row + roff, 1, rat_val)
            behavior_ws.cell(total_shock_header_row + roff, 2, rfid_val)
            behavior_ws.cell(total_shock_header_row + roff, 3, drug_val)
            behavior_ws.cell(total_shock_header_row + roff, 4, exp_val) #Experiment Group
            roff += 1


    # Fill exit day/code/notes for total shocks section
    max_row_behavior = behavior_ws.max_row
    for row_b in range(total_shock_header_row + 1, max_row_behavior + 1):
        b_rat = behavior_ws.cell(row=row_b, column=1).value
        if not b_rat:
            continue
        if b_rat in exit_data_map:
            day_v, code_v, notes_v, last_good_v = exit_data_map[b_rat] #Updated
            behavior_ws.cell(row=row_b, column=exit_day_col, value=day_v)
            behavior_ws.cell(row=row_b, column=exit_code_col, value=code_v)
            behavior_ws.cell(row=row_b, column=exit_notes_col, value=notes_v)
            behavior_ws.cell(row=row_b, column=last_good_col,  value=last_good_v) #Updated


    total_shock_rat_start = total_shock_header_row + 1
    total_shock_rat_ids = []
    for rr in range(total_shock_rat_start, behavior_ws.max_row + 1):
        val = behavior_ws.cell(rr, 1).value
        if val and val not in ["Rat", "Active Lever Presses", "Inactive Lever Presses", 
                               "breakpoint", "Total Shocks", "Reward # Got Shock 1"]:
            total_shock_rat_ids.append(val)


    shock_folder_path = os.path.join(excel_output_files, "SHOCK")
    if os.path.exists(shock_folder_path):
        for s_name in all_sessions_in_sheet:
            if not s_name:
                continue
            for f_name in os.listdir(shock_folder_path):
                if (f_name.startswith("MTF134") or f_name.startswith("BSB273")) \
                   and s_name in f_name and f_name.endswith("_output.xlsx") \
                   and cohort_pattern.search(f_name):


                    session_file_path = os.path.join(shock_folder_path, f_name)
                    session_wb = openpyxl.load_workbook(session_file_path)
                    session_ws = session_wb.active


                    total_shocks_row = None
                    subject_row_in_file = None
                    for row_cells in session_ws.iter_rows(min_col=1, max_col=1):
                        if row_cells[0].value == "Total Shocks":
                            total_shocks_row = row_cells[0].row
                        if row_cells[0].value == "Subject":
                            subject_row_in_file = row_cells[0].row


                    if total_shocks_row and subject_row_in_file:
                        total_shocks_data = [cell.value for cell in session_ws[total_shocks_row]]
                        subject_data = [cell.value for cell in session_ws[subject_row_in_file]]
                        for rat_idx, rat_id in enumerate(subject_data[1:], start=1):
                            if rat_id in total_shock_rat_ids:
                                for row_b in range(total_shock_rat_start, behavior_ws.max_row + 1):
                                    if behavior_ws.cell(row=row_b, column=1).value == rat_id:
                                        for c_b in range(FIRST_SESSION_COL, behavior_ws.max_column + 1):
                                            if behavior_ws.cell(row=2, column=c_b).value == s_name:
                                                behavior_ws.cell(row=row_b, column=c_b, value=total_shocks_data[rat_idx])
                                                break


    # -------------------------------------------------------------------------
    # REWARD # GOT SHOCK 1
    # -------------------------------------------------------------------------
    reward_shock1_row = find_or_create_section(behavior_ws, "Reward # Got Shock 1")
    reward_shock1_header_row = reward_shock1_row + 1
    if behavior_ws.cell(reward_shock1_header_row, 1).value != "Rat":
        behavior_ws.cell(reward_shock1_header_row, 1, "Rat")
        behavior_ws.cell(reward_shock1_header_row, 2, "RFID")
        behavior_ws.cell(reward_shock1_header_row, 3, "Drug Group")
        behavior_ws.cell(reward_shock1_header_row, 4, "Experiment Group") 
        roff = 1
        for (rat_val, rfid_val, drug_val, exp_val) in original_rat_rfid_dg:
            behavior_ws.cell(reward_shock1_header_row + roff, 1, rat_val)
            behavior_ws.cell(reward_shock1_header_row + roff, 2, rfid_val)
            behavior_ws.cell(reward_shock1_header_row + roff, 3, drug_val)
            behavior_ws.cell(reward_shock1_header_row + roff, 4, exp_val) #Experiment Group
            roff += 1


    # Fill exit day/code/notes for "Reward # Got Shock 1" section
    max_row_behavior = behavior_ws.max_row
    for row_b in range(reward_shock1_header_row + 1, max_row_behavior + 1):
        b_rat = behavior_ws.cell(row=row_b, column=1).value
        if not b_rat:
            continue
        if b_rat in exit_data_map:
            day_v, code_v, notes_v, last_good_v = exit_data_map[b_rat] #Updated
            behavior_ws.cell(row=row_b, column=exit_day_col, value=day_v) 
            behavior_ws.cell(row=row_b, column=exit_code_col, value=code_v)
            behavior_ws.cell(row=row_b, column=exit_notes_col, value=notes_v) 
            behavior_ws.cell(row=row_b, column=last_good_col,  value=last_good_v) #Updated


    reward_shock1_rat_start = reward_shock1_header_row + 1
    reward_shock1_rat_ids = []
    for rr in range(reward_shock1_rat_start, behavior_ws.max_row + 1):
        val = behavior_ws.cell(row=rr, column=1).value
        if val and val not in ["Rat", "Active Lever Presses", "Inactive Lever Presses", "breakpoint", 
                               "Total Shocks", "Reward # Got Shock 1"]:
            reward_shock1_rat_ids.append(val)


    # Now we only read from shock folder, row => "Reward # Got Shock 1"
    shock_folder_path = os.path.join(excel_output_files, "SHOCK")
    if os.path.exists(shock_folder_path):
        for s_name in all_sessions_in_sheet:
            if not s_name:
                continue
            for f_name in os.listdir(shock_folder_path):
                if (f_name.startswith("MTF134") or f_name.startswith("BSB273")) \
                   and s_name in f_name and f_name.endswith("_output.xlsx") \
                   and cohort_pattern.search(f_name):


                    session_file_path = os.path.join(shock_folder_path, f_name)
                    session_wb = openpyxl.load_workbook(session_file_path)
                    session_ws = session_wb.active


                    reward_shock1_data_row = None
                    subject_row_in_file = None
                    for row_cells in session_ws.iter_rows(min_col=1, max_col=1):
                        if row_cells[0].value == "Reward # Got Shock 1":
                            reward_shock1_data_row = row_cells[0].row
                        if row_cells[0].value == "Subject":
                            subject_row_in_file = row_cells[0].row


                    if reward_shock1_data_row and subject_row_in_file:
                        shock1_data = [cell.value for cell in session_ws[reward_shock1_data_row]]
                        subject_data = [cell.value for cell in session_ws[subject_row_in_file]]


                        for rat_idx, rat_id in enumerate(subject_data[1:], start=1):
                            if rat_id in reward_shock1_rat_ids:
                                for row_b in range(reward_shock1_rat_start, behavior_ws.max_row + 1):
                                    if behavior_ws.cell(row=row_b, column=1).value == rat_id:
                                        for c_b in range(FIRST_SESSION_COL, behavior_ws.max_column + 1):
                                            if behavior_ws.cell(row=2, column=c_b).value == s_name:
                                                behavior_ws.cell(row=row_b, column=c_b, value=shock1_data[rat_idx])
                                                break

    #Repeat Exit-tab info in every section - Updated
    section_markers = {
        "Rewards", "Active Lever Presses", "Inactive Lever Presses",
        "breakpoint", "Total Shocks", "Reward # Got Shock 1"
    }
    for hdr_row in range(1, behavior_ws.max_row + 1):
        # detect the “Rat / RFID / Drug Group” header of each section
        if (behavior_ws.cell(hdr_row, 1).value == "Rat"
                and behavior_ws.cell(hdr_row, 2).value == "RFID"
                and behavior_ws.cell(hdr_row, 3).value == "Drug Group"
                and behavior_ws.cell(hdr_row, 4).value == "Experiment Group"):

            r = hdr_row + 1     
            while r <= behavior_ws.max_row:
                first_val = behavior_ws.cell(r, 1).value
                # stop when we hit the next section (or blank)
                if first_val in section_markers or first_val in ("Rat", None):
                    break
                # fill Exit info
                if first_val in exit_data_map:                      
                    day_v, code_v, notes_v, last_good_v = exit_data_map[first_val]
                    behavior_ws.cell(r, exit_day_col,  day_v)
                    behavior_ws.cell(r, exit_code_col, code_v)
                    behavior_ws.cell(r, exit_notes_col, notes_v)
                    behavior_ws.cell(r, last_good_col, last_good_v)
                r += 1

    # -------------------------------------------------------------------------
    # 16) DAILY ISSUES & NOTES => Rows=3 & 4
    # -------------------------------------------------------------------------
    behavior_ws["A3"] = "Daily Issues"
    behavior_ws["A4"] = "Notes"


    max_col_now = behavior_ws.max_column
    for col in range(FIRST_SESSION_COL, max_col_now + 1):
        behavior_ws.cell(row=3, column=col, value=None)
        behavior_ws.cell(row=4, column=col, value=None)


    daily_issues_map = {}
    daily_notes_map = {}


    if os.path.exists(daily_issues_file_path):
        daily_issues_wb = openpyxl.load_workbook(daily_issues_file_path)
        daily_issues_ws = daily_issues_wb.active


        headers_map = {}
        for cidx, cell in enumerate(daily_issues_ws[1], start=1):
            hval = cell.value if cell.value else ""
            headers_map[hval] = cidx


        trial_id_col = headers_map.get("Trial ID")
        subject_col = headers_map.get("Subject")
        code_col = headers_map.get("Code")
        note_col = headers_map.get("Note")


        if not (trial_id_col and subject_col and code_col and note_col):
            print(f"Warning: 'Trial ID', 'Subject', 'Code', or 'Note' column missing in {daily_issues_file_name}")
        else:
            max_row_issues = daily_issues_ws.max_row
            for row_i in range(2, max_row_issues + 1):
                t_id = daily_issues_ws.cell(row=row_i, column=trial_id_col).value
                subj = daily_issues_ws.cell(row=row_i, column=subject_col).value
                codeval = daily_issues_ws.cell(row=row_i, column=code_col).value
                noteval = daily_issues_ws.cell(row=row_i, column=note_col).value


                if not t_id:
                    continue


                trial_upper = str(t_id).strip().upper()
                subject_str = str(subj).strip() if subj else ""
                code_str = str(codeval).strip() if codeval else ""
                note_str = str(noteval).strip() if noteval else ""


                # For Daily Issues => "Subject: Code ;"
                if subject_str or code_str:
                    combined_issue_str = f"{subject_str}: {code_str} ;"
                    daily_issues_map.setdefault(trial_upper, []).append(combined_issue_str)


                # For Notes => "Subject: Note ;"
                if note_str:
                    combined_note_str = f"{subject_str}: {note_str} ;"
                    daily_notes_map.setdefault(trial_upper, []).append(combined_note_str)


        session_col_map = {}
        for cidx in range(FIRST_SESSION_COL, behavior_ws.max_column + 1):
            sess_name = behavior_ws.cell(row=2, column=cidx).value
            if sess_name:
                session_col_map[sess_name.upper()] = cidx


        for trial_id_upper, issues_list in daily_issues_map.items():
            if trial_id_upper in session_col_map:
                col_idx = session_col_map[trial_id_upper]
                behavior_ws.cell(row=3, column=col_idx, value="\n".join(issues_list))


        for trial_id_upper, notes_list in daily_notes_map.items():
            if trial_id_upper in session_col_map:
                col_idx = session_col_map[trial_id_upper]
                behavior_ws.cell(row=4, column=col_idx, value="\n".join(notes_list))
    else:
        print(f"Warning: daily issues file not found at: {daily_issues_file_path}")


    # -------------------------------------------------------------------------
    # 17) SAVE (unsorted)
    # -------------------------------------------------------------------------
    # behavior_wb.save(behavior_sheet_path)
    behavior_wb.save(behavior_sheet_path)
    print(f"Data merged (for {cohort_str} cohort) with new 'Total Shocks' and 'Reward # Got Shock 1' sections. "
          f"Saved at: {behavior_sheet_path} (unsorted).")


    # -------------------------------------------------------------------------
    # 18) SORT COLUMNS BY DATE & REMOVE OLD SHEET
    # -------------------------------------------------------------------------
    sort_columns_by_date(behavior_wb, "Behavior Data", "Sorted Data")
    behavior_wb.save(behavior_sheet_path)
    #behavior_wb.save(behavior_sheet_path)
    print("Done! Columns sorted by date, old sheet removed, final file saved.")




if __name__ == "__main__":
    main()
