
from cmath import exp
from collections import defaultdict
from getoperant import GEToperant as getop 
from openpyxl import Workbook
import openpyxl
import os
import re
import pprint
import numpy as np
import shutil

'''
    The function has two parameters: lga_files and sha_files
    Both parameters are lists of files of its respective experiment type that are to be processed
    
'''

def lga_sha_extract_refactor(extract_filepaths):

    print('LGA_SHA_EXTRACTING')

    # TODO: ERROR AS THE MEDPC FILES MAY HAVE A LINE ABOUT BACKUP METADATA
    # 2nd PR file in OXYCODONE

    # Keeping track of cohorts and experiements over 2000 presses
    cohort_list_error_presses = {}

    for tuple_name_path in extract_filepaths:

        filename = tuple_name_path[0]

        filename_path = tuple_name_path[1]

        filename_lower = filename.lower()

        if 'oxy' in filename_lower and 'lga' in filename_lower:
            drug_folder, exp_folder = 'OXYCODONE', 'LGA'
        elif 'oxy' in filename_lower and 'sha' in filename_lower:
            drug_folder, exp_folder = 'OXYCODONE', 'SHA'
        # Case for COCAINE, lga as cocaine has no identifier in filename
        elif 'lga' in filename_lower:
            drug_folder, exp_folder = 'COCAINE', 'LGA'
        else:
            drug_folder, exp_folder = 'COCAINE', 'SHA'        
        
        # Uses the profile from GEToperantShA.xlsx 
        profile_file_name = 'C:/Users/georg/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/data_automation/getop_profiles/lga_sha_getop_profile.xlsx'
        
        # Testing on personal computer
        # profile_file_name = '/Users/haltand/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/data_automation/getop_profiles/lga_sha_getop_profile.xlsx'


        # Arguments for GEToperant function
        medpc_test_files = [filename_path]
        output_file_name = f'{filename}_output.xlsx'

        output_filepath = f'C:/Users/georg/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/excel_output_files/{drug_folder}/{exp_folder}/{output_file_name}'
        azure_output_filepath = f'C:/Users/georg/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/azure_excel_output_files/{drug_folder}/{exp_folder}/{output_file_name}'

        # Testing on personal computer
        # output_filepath = f'/Users/haltand/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/excel_output_files/{drug_folder}/{exp_folder}/{output_file_name}'
        # azure_output_filepath = f'/Users/haltand/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/azure_excel_output_files/{drug_folder}/{exp_folder}/{output_file_name}'

        # TODO: Currently used with manual_file_extraction as the filepath for the user changes
        # output_filepath = f'/Users/haltand/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/excel_output_files/{drug_folder}/{exp_folder}/{output_file_name}'


        # Remove files if they exist to rerun script with new output
        if os.path.exists(output_filepath):
            os.remove(output_filepath)
            print('Old file deleted')
        else:
            print('No file to delete', output_filepath)

        if os.path.exists(azure_output_filepath):
            os.remove(azure_output_filepath)
            print('Old file deleted')
        else:
            print('No file to delete', azure_output_filepath)

        with open(f"{filename_path}", "r+") as fp:

            lines = fp.readlines()

            if lines[-1] == '\x1a' or lines[-1] == '.':
            
                fp.writelines(lines[:-1])

            fp.close()

        # Figure out format to get more precise data into the spreadsheet
        getop.GEToperant(profile_file_name, medpc_test_files, output_filepath)
        getop.GEToperant(profile_file_name, medpc_test_files, azure_output_filepath)
        
        print('New file created at:', output_filepath)
        print('New file created at:', azure_output_filepath)

        wb = openpyxl.load_workbook(output_filepath)
        wb_azure = openpyxl.load_workbook(azure_output_filepath)

        ws = wb.active 
        ws_azure = wb_azure.active

        dict_subject_timeout_presses = list()

        # Used to pad matrix with 0 to append to spreadsheet
        max_length_timeout_presses = 0
        subject_over_2000_presses = []

        # Itearte through the spreadsheet by columns, rows
        for col_num in range(1, len(list(ws.iter_cols()))):

            active_presses = []
            reward_presses = []
            subject_name = ""
            over_2000_presses = False
            
            for row in ws.iter_rows():

                cell_value = row[col_num].value

                if row[0].value == 'Subject':
                    subject_name = cell_value

                # Handle the case that active lever presses are above 2000 and causes error
                # Keep track of the information
                if row[0].value == 'Active Lever Presses' and int(cell_value) > 2000:
                    subject_over_2000_presses.append(subject_name)
                    over_2000_presses = True

                # Regex match on Active # but not on 'Active Lever Presses'
                if re.match(r"Active \d+", str(row[0].value)) and cell_value != 0:
                    active_presses.append(cell_value)

                # Regex match on Reward # but not on 'Reward'
                if re.match(r"Reward \d+", str(row[0].value)) and cell_value != 0:
                    reward_presses.append(cell_value)
            
            # print('Active presses list is:', active_presses)
            # print('Reward presses list:', reward_presses)
            # Timeout presses = active presses - reward presses
            if not over_2000_presses:
                for num in reward_presses:
                    active_presses.remove(num) 

                dict_subject_timeout_presses.append(active_presses)
            # Case when we ignore experiments with over 2000 active lever presses
            else:
                dict_subject_timeout_presses.append([])
                # subject_over_2000_presses.append(subject_name)
        
        cohort_list_error_presses[filename] = subject_over_2000_presses

        # Gets the max length of the timeout presses to format the data better below (into m x m matrix)
        max_length_timeout_presses = max(len(lst) for lst in dict_subject_timeout_presses)
        
        # print(f'Largest amount of timeout presses: {max_length_timeout_presses}')

        # Prints out the times of timeout presses for each experiment
        pp = pprint.PrettyPrinter()
        # for key, value in dict_subject_timeout_presses.items():
        #     pp.pprint(f'{key} had {len(value)} timeout presses: {value}')

        # Format data into m x m matrix to modify the spreadsheet with new rows
        for index in range(len(dict_subject_timeout_presses)):
            temp_col = dict_subject_timeout_presses[index]
            while len(temp_col) < max_length_timeout_presses:
                temp_col.append(0)
            dict_subject_timeout_presses[index] = temp_col

        # Print formatted data by experiment
        # for k, v in dict_subject_timeout_presses.items():
        #     pp.pprint(f'{k}: {v}')

        # print('Number of columns in this file:', len(dict_subject_timeout_presses))

        np_matrix = []

        # Store experiment outputs into a 2d list
        for exp_col in dict_subject_timeout_presses:
            np_matrix.append(exp_col)

        # Formatted columns into a matrix
        matrix_output = np.column_stack(np_matrix)

        for i in range(0, max_length_timeout_presses):
            ws.append(list(np.append([f'Timeout Press {i+1}'], matrix_output[i])))

        # Save the changes
        wb.save(output_filepath)
        wb.save(azure_output_filepath)
        
    print(f'Cohort list of error presses: {cohort_list_error_presses}')