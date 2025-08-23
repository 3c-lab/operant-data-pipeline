from getoperant import GEToperant as getop 
from openpyxl import Workbook
# from typing import *
import openpyxl
import os

def shock_extract_refactor(extract_filepaths):

    print('SHOCK_EXTRACTING')

    for tuple_name_path in extract_filepaths:

            filename = tuple_name_path[0]

            filename_path = tuple_name_path[1]

            filename_lower = filename.lower()

            if 'oxy' in filename_lower:
                drug_folder, exp_folder = 'OXYCODONE', 'SHOCK'
            else:
                drug_folder, exp_folder = 'COCAINE', 'SHOCK'
            
            # Uses the profile from GEToperantShA.xlsx 
            profile_file_name = 'C:/Users/georg/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/data_automation/getop_profiles/shock_getop_profile.xlsx'
            
            # Testing on personal computer
            # profile_file_name = '/Users/haltand/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/data_automation/getop_profiles/shock_getop_profile.xlsx'      

            # Arguments for GEToperant function
            medpc_test_files = [filename_path]
            output_file_name = f'{filename}_output.xlsx'

            # TODO: Create a subdirectory for each individual output
            output_filepath = f'C:/Users/georg/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/excel_output_files/{drug_folder}/{exp_folder}/{output_file_name}'
            azure_output_filepath = f'C:/Users/georg/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/azure_excel_output_files/{drug_folder}/{exp_folder}/{output_file_name}'

            # TODO: Currently used with manual_file_extraction as the filepath for the user changes
            # Testing on personal computer
            # output_filepath = f'/Users/haltand/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/excel_output_files/{drug_folder}/{exp_folder}/{output_file_name}'
            # azure_output_filepath = f'/Users/haltand/George Lab Dropbox/George_Lab/Experiments/DataStream/DataSource/azure_excel_output_files/{drug_folder}/{exp_folder}/{output_file_name}'

            # Remove files if they exist to rerun script with new output
            if os.path.exists(output_filepath):
                os.remove(output_filepath)
                print('Old file deleted')
            else:
                print('No file to delete', output_filepath)

            if os.path.exists(azure_output_filepath):
                os.remove(azure_output_filepath)
                print('Old file deleted')
            else:
                print('No file to delete', azure_output_filepath)

            # Figure out format to get more precise data into the spreadsheet
            getop.GEToperant(profile_file_name, medpc_test_files, output_filepath)
            getop.GEToperant(profile_file_name, medpc_test_files, azure_output_filepath)

            print('New file created at:', output_filepath)
            print('New file created at:', azure_output_filepath)

            wb = openpyxl.load_workbook(output_filepath)
            wb_azure = openpyxl.load_workbook(azure_output_filepath)
            ws = wb.active
            ws_azure = wb_azure.active

            rewards_after_first_shock = ['Rewards After First Shock']

            for col_num in range(1, len(list(ws.iter_cols()))): 
                total_reward = 0
                reward_index_first_shock = 0

                for row in ws.iter_rows():

                    if row[0].value == "Total Reward":
                        total_reward = row[col_num].value
                    if row[0].value == "Reward # Got Shock 1":
                        reward_index_first_shock = row[col_num].value

                rewards_after_first_shock.append(total_reward - reward_index_first_shock)

            # Possibly be redundant - can just save the workout with a different file path (output_filepath and azure_output_filepath)
            ws.append(rewards_after_first_shock)
            ws_azure.append(rewards_after_first_shock)
            wb.save(output_filepath)
            wb.save(azure_output_filepath)